# -*- coding:utf-8 -*-
__author__ = 'shijin'
from openpyxl import Workbook  # 新增
from openpyxl import load_workbook  # 加载已有
import collections
from row_object import RowObject, RowStatus


# work_book = load_workbook('./vid-20210214.xlsx')
# # 获取sheet页name,返回数组
# print(work_book.sheetnames)
# # 获取sheet页数据对象
# sheet_data = work_book["listing"]
# print(dir(sheet_data))
# print(type(sheet_data))
# print(sheet_data.title)
# # print(sheet_data.tables)
# # print(sheet_data.path)
# # print(sheet_data.parent)
# # print(sheet_data.iter_rows().__next__())
# print(sheet_data.iter_rows(min_row=1, max_row=5, values_only=True).__next__())


# print(sheet_data)
# print(sheet_data.max_row)
# print(sheet_data.max_column)
# 获取单元格内的值，cell内的索引从1开始
# print(sheet_data.cell(1, 1).value)
# print(sheet_data.cell(1, 2).value)
# print(sheet_data.cell(1, 3).value)


# 修改数据并保存
# sheet_data.cell(1, 1).value = "哈哈哈"
# work_book.save('./123.xlsx')


class HandleXLSX(object):
    """
    Class for handling Excel file
    """

    def __init__(self, file_name, sheet_name=None, **kwargs):
        """
        :param file_name: The excel file name
        :param sheet_name: The sheet name in excel file
        :param kwargs:
        """
        self._file_name = file_name
        self.work_book = self._file_name
        self.sheet = sheet_name
        self.row_title_list = []

    @property
    def work_book(self):
        return self._work_book

    @work_book.setter
    def work_book(self, file_name):
        self._work_book = load_workbook(file_name)

    @property
    def sheet(self):
        return self._sheet

    @sheet.setter
    def sheet(self, sheet_name=None):
        if sheet_name:
            self._sheet = self._work_book[str(sheet_name).strip()]
        else:
            self._sheet = self._work_book.active

    @property
    def max_column(self):
        return self._sheet.max_column

    @property
    def max_row(self):
        return self._sheet.max_row

    def get_rows_start_end(self, start_point=None, end_point=None, **kwargs):
        """
        Check the start and end points of rows
        :param start_point: >= 2, 1 is the column name row.
        :param end_point: >= start_point and <= max rows
        :param kwargs:
        :return: start and end points after verification
        """
        if start_point is None:
            _start_point = 2
        elif isinstance(start_point, int):
            _start_point = start_point if start_point >= 2 else 2
        elif isinstance(start_point, str):
            assert start_point.strip().isdigit(), "Parameter start_point must be int"
            _start_point = int(start_point) if int(start_point) >= 2 else 2
        else:
            raise ValueError(f'{start_point} is invalid')

        if end_point is None:
            _end_point = self.max_row
        elif isinstance(end_point, int):
            _end_point = end_point if end_point <= self.max_row else self.max_row
        elif isinstance(end_point, str):
            assert end_point.strip().isdigit(), "Parameter end_point must be int"
            _end_point = int(end_point) if int(end_point) <= self.max_row else self.max_row
        else:
            raise ValueError(f'{end_point} is invalid')

        assert _start_point <= _end_point, "start_point is not less than 2, " \
                                           "end_point must be more than or equal to start_point"
        return _start_point, _end_point

    def get_sheet_iterator(self, sheet_name=None, start_point=None, end_point=None, **kwargs):
        """
        :param sheet_name: sheet name
        :param start_point: the start row number of the sheet
        :param end_point: the end row number of the sheet
        :param kwargs:
        :return:
        """
        if sheet_name:
            self.sheet = sheet_name

        column_name_list = self.get_column_name_list(sheet_name, **kwargs)
        _sheet_iter_rows = self.sheet.iter_rows(min_row=start_point, max_row=end_point, min_col=1,
                                                max_col=len(column_name_list))
        return _sheet_iter_rows

    def generator_rows_value_iterator(self, sheet_name=None, start_point=None, end_point=None, **kwargs):
        """
        :param sheet_name: sheet name
        :param start_point: the start row number of the table
        :param end_point: the end row number of the table
        :param kwargs:
        :return:
        """
        column_name_list = self.get_column_name_list(sheet_name, **kwargs)
        _start_point, _end_point = self.get_rows_start_end(start_point, end_point, **kwargs)
        _sheet_iter_rows = self.get_sheet_iterator(sheet_name, start_point=_start_point, end_point=_end_point, **kwargs)

        for row in _sheet_iter_rows:
            row_value_list = [cell.value for cell in row]

            _row_dict = self.create_row_dict(column_name_list, row_value_list)
            if _row_dict:
                yield _row_dict

    def generate_row_object_iterator(self, check_file, sheet_name=None, start_point=None, end_point=None, **kwargs):
        """
        :param check_file:
        :param sheet_name: sheet name
        :param start_point: the start row number of the sheet
        :param end_point: the end row number of the sheet
        :param kwargs:
        :return: row object iterator
        """
        column_name_list = self.get_column_name_list(sheet_name, **kwargs)
        _start_point, _end_point = self.get_rows_start_end(start_point, end_point, **kwargs)
        _sheet_iter_rows = self.get_sheet_iterator(sheet_name, start_point=_start_point, end_point=_end_point, **kwargs)

        for row_number, row in enumerate(_sheet_iter_rows, start=_start_point):
            row_value_list = [cell.value for cell in row]
            if row_value_list:
                row_object = RowObject()
                row_object.file_name = self._file_name
                row_object.sheet_name = self.sheet.title
                row_object.column_name = column_name_list
                row_object.row_value = row_value_list
                row_object.column_value = dict(zip(column_name_list, row_value_list))
                row_object.position = row_number
                row_object.status = RowStatus.INITIAL.value if check_file is False else RowStatus.CHECK.value
                yield row_object

    def get_column_name_list(self, sheet_name=None, **kwargs):
        """
        Get column name in sheet, it is the first row
        :param sheet_name:
        :param kwargs:
        :return:
        """
        if sheet_name:
            self.sheet = sheet_name
        _column_name_list = self.sheet.iter_rows(values_only=True).__next__()
        return list(_column_name_list)

    def write_sheet_rows_value(self, sheet_name, values, auto_save=False, **kwargs):
        """
        Write cell value in sheet
        :param sheet_name:
        :param values:
        :param auto_save:
        :param kwargs:
        :return:
        """
        self.sheet = sheet_name

        assert isinstance(values, (list, tuple)), "values must be list or tuple"
        if len(values) == 3 and isinstance(values[0], int) and isinstance(values[1], int):
            # For example self.sheet.cell(1, 5, "哈哈哈")
            self.sheet.cell(*values)
        else:
            for cell in values:
                assert isinstance(cell, (list, tuple)) and len(cell) == 3, "parameters must be x, y, value"
                # x = cell[0]
                # y = cell[1]
                # value = cell[2]
                self.sheet.cell(*cell)

        if auto_save is True:
            self.save()

    def save(self, file_name=None):
        """
        save Excel File
        :param file_name: output file name
        :return:
        """
        # self._work_book.save('./123.xlsx')
        _file_name = str(file_name).strip() if file_name else self._file_name
        self._work_book.save(_file_name)

    def close(self):
        self._work_book.close()

    def create_row_dict(self, row_title_list, row_value_list, **kwargs):
        """
        create row data structure
        :param row_title_list: title list of table
        :param row_value_list: each row list of table
        :return: order dict
        """
        _row_dict = collections.OrderedDict()
        if row_title_list and row_value_list:
            for index, key in enumerate(row_title_list):
                _row_dict[key] = row_value_list[index]
        return _row_dict

    def prepare_write_value(self, row_dict, result):
        assert isinstance(row_dict, dict), "row_dict must be OrderedDict"
        x = int(row_dict.get('sn')) + 1
        y = list(row_dict.keys()).index('result')
        value = result
        return x, y, value

    def prepare_write_result(self, download_result, result='Success'):
        if isinstance(download_result, (list, tuple)):
            for _row_dict in download_result:
                self.prepare_write_value(_row_dict, result)
        else:
            self.prepare_write_value(download_result, result)


if __name__ == "__main__":
    print("###" * 30)
    filename = 'vid-20210214.xlsx'
    xlsl = HandleXLSX(filename)
    rows = xlsl.generator_rows_value_iterator(sheet_name='listing', start_point=5, end_point=10)

    for i in rows:
        print("row: ", i)

    print("***" * 30)
    rows = xlsl.generate_row_object_iterator(sheet_name='listing', start_point=5, end_point=10)

    for j in rows:
        print("row: ", j, j.__dict__)
